#!/usr/bin/env python3
"""B题 — 最终汇总：生成所有交付文件 (xlsx, MONITOR, paper)"""
import json, os, shutil, copy
from collections import defaultdict

OUT = 'outputs/b-tube-cut'
os.makedirs(OUT, exist_ok=True)

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

TEMPLATE_DIR = 'fixtures/t3/tube_cut_b2026/raw/B题 附件/B题 结果'

def load_sol(name):
    with open(os.path.join(OUT, name)) as f:
        return json.load(f)

def compress_seq(seq):
    if not seq: return ''
    parts = []
    i = 0
    while i < len(seq):
        g = seq[i]; count = 1
        while i+count < len(seq) and seq[i+count] == g:
            count += 1
        parts.append(f'G{g}×{count}')
        i += count
    return '|'.join(parts)

def load_savings():
    with open(os.path.join(OUT, 'cocut_savings.json')) as f:
        return json.load(f)

def write_result_xlsx(stocks_data, result_num, summary_data, has_cocut=False):
    """Write a result xlsx from template. stocks_data = list of stock dicts."""
    template_path = os.path.join(TEMPLATE_DIR, f'result{result_num}.xlsx')
    out_path = os.path.join(OUT, f'result{result_num}.xlsx')
    
    if os.path.exists(template_path):
        shutil.copy(template_path, out_path)
        wb = openpyxl.load_workbook(out_path)
    else:
        wb = openpyxl.Workbook()
    
    # ── Sheet 1: 下料方案 ──
    ws_name = '下料方案'
    if ws_name in wb.sheetnames:
        ws = wb[ws_name]
    else:
        ws = wb.active
        ws.title = ws_name
    
    # Find header row or create one
    data_start = 1
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and '母材编号' in str(val):
            data_start = row
            break
    if data_start == 1:
        data_start = ws.max_row + 1 if ws.max_row > 1 else 1
    
    # Clear old data below header
    for row in range(data_start + 1, ws.max_row + 1):
        for col in range(1, 7):
            ws.cell(row=row, column=col).value = None
    
    # Write data
    header_font = Font(bold=True)
    for i, s in enumerate(stocks_data):
        row = data_start + 1 + i
        seq_str = compress_seq(s['sequence'])
        stock_len = s.get('stock_len', 0)
        raw_used = s.get('used_len', 0)
        benefit = s.get('cocut_benefit', 0)
        effective = raw_used - benefit
        waste = stock_len - effective
        util = effective / stock_len if stock_len > 0 else 0
        
        ws.cell(row=row, column=1, value=s.get('stock_id', f'M{i+1}'))
        ws.cell(row=row, column=2, value=round(stock_len, 1))
        ws.cell(row=row, column=3, value=seq_str)
        ws.cell(row=row, column=4, value=round(effective, 1))
        ws.cell(row=row, column=5, value=round(waste, 1))
        ws.cell(row=row, column=6, value=round(util, 4))
    
    # ── Sheet 2: 拼接方式摘要表 (if applicable) ──
    if has_cocut:
        ws2_name = '拼接方式摘要表'
        if ws2_name in wb.sheetnames:
            ws2 = wb[ws2_name]
        else:
            ws2 = wb.create_sheet(ws2_name)
        
        # Find or create header
        j_start = 1
        for row in range(1, ws2.max_row + 1):
            val = ws2.cell(row=row, column=1).value
            if val and '母材编号' in str(val):
                j_start = row
                break
        if j_start == 1:
            j_start = ws2.max_row + 1 if ws2.max_row > 1 else 1
        
        # Clear old data
        for row in range(j_start + 1, ws2.max_row + 1):
            for col in range(1, 9):
                ws2.cell(row=row, column=col).value = None
        
        jr = j_start + 1
        savings = load_savings()
        
        for s in stocks_data:
            seq = s['sequence']
            stock_id = s.get('stock_id', '?')
            if len(seq) < 2:
                continue
            
            # Build blocks
            blocks = []
            i = 0
            while i < len(seq):
                g = seq[i]; count = 1
                while i+count < len(seq) and seq[i+count] == g:
                    count += 1
                blocks.append((g, count))
                i += count
            
            # Internal joints
            for g, count in blocks:
                if count > 1:
                    best_mode = max(['LL','LR','RL','RR'], key=lambda m: savings.get(f'G{g}-G{g}',{}).get(m,0))
                    bp = savings.get(f'G{g}-G{g}', {}).get(best_mode, 0)
                    ws2.cell(row=jr, column=1, value=stock_id)
                    ws2.cell(row=jr, column=2, value='内部拼接')
                    ws2.cell(row=jr, column=3, value=f'G{g}×{count}')
                    ws2.cell(row=jr, column=4, value=f'G{g}×{count}')
                    ws2.cell(row=jr, column=5, value=best_mode)
                    ws2.cell(row=jr, column=6, value=count-1)
                    ws2.cell(row=jr, column=7, value=round(bp, 3))
                    ws2.cell(row=jr, column=8, value=round(bp * (count-1), 3))
                    jr += 1
            
            # Inter-block joints
            for bi in range(len(blocks) - 1):
                g1, c1 = blocks[bi]
                g2, c2 = blocks[bi+1]
                best_mode = max(['LL','LR','RL','RR'], key=lambda m: savings.get(f'G{g1}-G{g2}',{}).get(m,0))
                bp = savings.get(f'G{g1}-G{g2}', {}).get(best_mode, 0)
                ws2.cell(row=jr, column=1, value=stock_id)
                ws2.cell(row=jr, column=2, value='块间拼接')
                ws2.cell(row=jr, column=3, value=f'G{g1}×{c1}')
                ws2.cell(row=jr, column=4, value=f'G{g2}×{c2}')
                ws2.cell(row=jr, column=5, value=best_mode)
                ws2.cell(row=jr, column=6, value=1)
                ws2.cell(row=jr, column=7, value=round(bp, 3))
                ws2.cell(row=jr, column=8, value=round(bp, 3))
                jr += 1
    
    # ── Sheet 3: 汇总 ──
    ws3_name = '汇总'
    if ws3_name in wb.sheetnames:
        ws3 = wb[ws3_name]
    else:
        ws3 = wb.create_sheet(ws3_name)
    
    max3 = ws3.max_row
    sr = max3 + 2 if max3 > 1 else 1
    items = [
        ('总母材长度(mm)', summary_data.get('total_stock_length', 0)),
        ('总切换次数', summary_data.get('total_switches', 0)),
        ('总利用率', round(summary_data.get('total_stock_length', 0) > 0 and 
                   sum(s.get('used_len',0)-(s.get('cocut_benefit',0)) for s in stocks_data) / 
                   max(sum(s['stock_len'] for s in stocks_data),1), 4)),
    ]
    if has_cocut:
        items.append(('总共切收益(mm)', round(summary_data.get('total_cocut_benefit', 0), 2)))
    
    for idx, (label, val) in enumerate(items):
        ws3.cell(row=sr+idx, column=1, value=label)
        ws3.cell(row=sr+idx, column=2, value=val)
    
    wb.save(out_path)
    print(f"  result{result_num}.xlsx: {len(stocks_data)} stocks written")

# ── MAIN: Generate all xlsx ──
print("=" * 60)
print("Generating result xlsx files...")
print("=" * 60)

q1 = load_sol('q1-solution.json')
q2 = load_sol('q2-solution.json')
q3 = load_sol('q3-solution.json')
q4 = load_sol('q4-solution.json')

write_result_xlsx(q1['stocks'], 1, q1, has_cocut=False)
write_result_xlsx(q2['stocks'], 2, q2, has_cocut=True)
write_result_xlsx(q3['stocks'], 3, q3, has_cocut=True)

# Q4: use batch_results with proper stock_ids
# For Q4, gather all stocks from batch results
q4_all_stocks = []
for bn in [1,2,3]:
    br = q4['batch_results'].get(str(bn), q4['batch_results'].get(bn))
    if br:
        for idx, s in enumerate(br.get('stocks', [])):
            s2 = dict(s)
            s2['stock_id'] = f'B{bn}_M{idx+1}'
            q4_all_stocks.append(s2)

# Also include inventory stocks
for s in q4.get('all_stocks', []):
    if s not in q4_all_stocks:
        q4_all_stocks.append(s)

write_result_xlsx(q4_all_stocks, 4, q4, has_cocut=True)

print("\nDone! All xlsx files written to outputs/b-tube-cut/")
