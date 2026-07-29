#!/usr/bin/env python3
"""B题 — 生成 result1~4.xlsx，从模板复制并填写数据"""
import json, os, shutil, copy
from collections import defaultdict

OUT = 'outputs/b-tube-cut'
TEMPLATE_DIR = 'fixtures/t3/tube_cut_b2026/raw/B题 附件/B题 结果'
os.makedirs(OUT, exist_ok=True)

# Try to import openpyxl
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed. Will generate CSV alternatives.")

# Load solution data
def load_sol(name):
    with open(os.path.join(OUT, name)) as f:
        return json.load(f)

q1 = load_sol('q1-solution.json')
q2 = load_sol('q2-solution.json')
q3 = load_sol('q3-solution.json')
q4 = load_sol('q4-solution.json')

with open(os.path.join(OUT, 'axial_lengths.json')) as f:
    axial = json.load(f)

def compress_seq(seq):
    if not seq:
        return ''
    parts = []
    i = 0
    while i < len(seq):
        g = seq[i]; count = 1
        while i+count < len(seq) and seq[i+count] == g:
            count += 1
        parts.append(f'G{g}×{count}')
        i += count
    return '|'.join(parts)

def write_result_xlsx(sol, result_num, has_cocut=False, is_q4=False):
    """Write result to xlsx from template."""
    template_path = os.path.join(TEMPLATE_DIR, f'result{result_num}.xlsx')
    out_path = os.path.join(OUT, f'result{result_num}.xlsx')
    
    if not HAS_OPENPYXL:
        # Fallback: copy template and note it needs filling
        if os.path.exists(template_path):
            shutil.copy(template_path, out_path)
        print(f"  result{result_num}.xlsx: copied template (openpyxl not available, needs manual fill)")
        return
    
    # Copy template
    if os.path.exists(template_path):
        shutil.copy(template_path, out_path)
        wb = openpyxl.load_workbook(out_path)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '下料方案'
    
    # ── Sheet 1: 下料方案 (Cutting Plan) ──
    if '下料方案' in wb.sheetnames:
        ws = wb['下料方案']
    elif len(wb.sheetnames) > 0:
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb.active
    
    # Find the data area and clear it, then fill
    # Template structure: row headers + data rows
    # We'll find the data starting row and write there
    
    # Clear old data (if any) - just append after template headers
    # For simplicity, write to a new area
    stocks_to_write = sol.get('stocks', [])
    if is_q4:
        # For Q4, get all stocks from all batches
        stocks_to_write = sol.get('all_stocks', [])
        if not stocks_to_write:
            for bn in [1,2,3]:
                br = sol.get('batch_results', {}).get(bn)
                if br:
                    stocks_to_write.extend(br.get('stocks', []))
    
    # Find last row with data
    max_row = ws.max_row
    data_start = max_row + 2  # Start after existing data
    
    # Write headers
    headers = ['母材编号(M_ID)', '母材长度(mm)', '工件块序列', '轴向占用总长度(mm)', '剩余长度(mm)', '母材利用率']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for i, s in enumerate(stocks_to_write):
        row = data_start + 1 + i
        seq_str = compress_seq(s['sequence'])
        stock_len = s.get('stock_len', 0)
        used = s.get('used_effective', s.get('used_len', 0))
        waste = stock_len - used
        util = used / stock_len if stock_len > 0 else 0
        
        stock_id = s.get('stock_id', f'M{i+1}')
        ws.cell(row=row, column=1, value=stock_id)
        ws.cell(row=row, column=2, value=round(stock_len, 1))
        ws.cell(row=row, column=3, value=seq_str)
        ws.cell(row=row, column=4, value=round(used, 1))
        ws.cell(row=row, column=5, value=round(waste, 1))
        ws.cell(row=row, column=6, value=round(util, 4))
    
    # ── Sheet 2: 拼接方式摘要表 (if has co-cutting) ──
    if has_cocut and '拼接方式摘要表' in wb.sheetnames:
        ws2 = wb['拼接方式摘要表']
    elif has_cocut:
        ws2 = wb.create_sheet('拼接方式摘要表')
    else:
        ws2 = None
    
    if ws2 and has_cocut:
        # Write joint details
        joint_headers = ['母材编号M_ID', '拼接类型', '前工件块', '后工件块', '拼接方式', '拼接次数', '单次共切收益(mm)', '共切收益小计(mm)']
        max_row2 = ws2.max_row
        j_start = max_row2 + 2
        
        for col, h in enumerate(joint_headers, 1):
            cell = ws2.cell(row=j_start, column=col, value=h)
            cell.font = Font(bold=True)
        
        jr = j_start + 1
        for s in stocks_to_write:
            joints = s.get('joints', [])
            if not joints:
                continue
            seq = s['sequence']
            stock_id = s.get('stock_id', '?')
            
            # Group joints by type (internal vs inter-block)
            # Compute compressed blocks
            blocks = []
            i = 0
            while i < len(seq):
                g = seq[i]; count = 1
                while i+count < len(seq) and seq[i+count] == g:
                    count += 1
                blocks.append((g, count))
                i += count
            
            # Internal joints within each block
            joint_idx = 0
            for g, count in blocks:
                if count > 1:
                    # n pieces → n-1 internal joints
                    # Determine best mode for Gg-Gg
                    with open(os.path.join(OUT, 'cocut_savings.json')) as f:
                        sav = json.load(f)
                    best_mode = max(['LL','LR','RL','RR'], key=lambda m: sav.get(f'G{g}-G{g}',{}).get(m,0))
                    benefit_per = sav.get(f'G{g}-G{g}', {}).get(best_mode, 0)
                    ws2.cell(row=jr, column=1, value=stock_id)
                    ws2.cell(row=jr, column=2, value='内部拼接')
                    ws2.cell(row=jr, column=3, value=f'G{g}×{count}')
                    ws2.cell(row=jr, column=4, value=f'G{g}×{count}')
                    ws2.cell(row=jr, column=5, value=best_mode)
                    ws2.cell(row=jr, column=6, value=count-1)
                    ws2.cell(row=jr, column=7, value=round(benefit_per, 3))
                    ws2.cell(row=jr, column=8, value=round(benefit_per * (count-1), 3))
                    jr += 1
            
            # Inter-block joints
            for bi in range(len(blocks) - 1):
                g1, c1 = blocks[bi]
                g2, c2 = blocks[bi+1]
                with open(os.path.join(OUT, 'cocut_savings.json')) as f:
                    sav = json.load(f)
                best_mode = max(['LL','LR','RL','RR'], key=lambda m: sav.get(f'G{g1}-G{g2}',{}).get(m,0))
                benefit = sav.get(f'G{g1}-G{g2}', {}).get(best_mode, 0)
                ws2.cell(row=jr, column=1, value=stock_id)
                ws2.cell(row=jr, column=2, value='块间拼接')
                ws2.cell(row=jr, column=3, value=f'G{g1}×{c1}')
                ws2.cell(row=jr, column=4, value=f'G{g2}×{c2}')
                ws2.cell(row=jr, column=5, value=best_mode)
                ws2.cell(row=jr, column=6, value=1)
                ws2.cell(row=jr, column=7, value=round(benefit, 3))
                ws2.cell(row=jr, column=8, value=round(benefit, 3))
                jr += 1
    
    # ── Sheet 3: 汇总 (Summary) ──
    if '汇总' in wb.sheetnames:
        ws3 = wb['汇总']
        # Add summary data
        max_row3 = ws3.max_row
        sr = max_row3 + 2
        ws3.cell(row=sr, column=1, value='总母材长度(mm)')
        ws3.cell(row=sr, column=2, value=sol.get('total_stock_length', 0))
        ws3.cell(row=sr+1, column=1, value='总切换次数')
        ws3.cell(row=sr+1, column=2, value=sol.get('total_switches', 0))
        if has_cocut:
            ws3.cell(row=sr+2, column=1, value='总共切收益(mm)')
            ws3.cell(row=sr+2, column=2, value=round(sol.get('total_cocut_benefit', 0), 2))
    
    wb.save(out_path)
    print(f"  result{result_num}.xlsx: written with {len(stocks_to_write)} stocks")

# ── Generate all result xlsx ──
print("Generating result xlsx files...")
write_result_xlsx(q1, 1, has_cocut=False)
write_result_xlsx(q2, 2, has_cocut=True)
write_result_xlsx(q3, 3, has_cocut=True)
write_result_xlsx(q4, 4, has_cocut=True, is_q4=True)
print("Done!")
