#!/usr/bin/env python3
"""Fixed tube-cut solver for 2026 MCM B (异形圆管下料).

Bugs fixed vs Pi draft:
1) Axial length = PCA first-axis span (tube axis), NOT Z_max-Z_min (~40mm disk).
2) Co-cut Δ = l_i+l_j - L_ab via end-envelope nesting + rotation search.
3) Cutting-stock: genuine mixed stock, exact orientation DP, seeded ALNS and beam search.
4) Export result1-4.xlsx + consistent DONE metrics from one source of truth.
"""
from __future__ import annotations

import json
import argparse
import hashlib
import shutil
from copy import copy, deepcopy
from pathlib import Path

import numpy as np
import pandas as pd

from tube_optimization import (
    counts_from_bins,
    homogeneous_block_stock_patterns,
    joint_relaxation_stock_lower_bound,
    minimize_type_splits_for_fixed_stock,
    mixed_stock_patterns,
    optimize_fixed_assignments,
    optimize_joint_bins,
    solve_multibatch_beam,
    summarize_bins,
)

ROOT = Path(__file__).resolve().parents[1]
CSV_DIR = (
    ROOT
    / "fixtures/t3/tube_cut_b2026/raw/B题 附件/B题 数据/附件1_10种工件"
)
BATCH_XLSX = (
    ROOT
    / "fixtures/t3/tube_cut_b2026/raw/B题 附件/B题 数据/附件2_三批次工件需求数据.xlsx"
)
TMPL_DIR = ROOT / "fixtures/t3/tube_cut_b2026/raw/B题 附件/B题 结果"
OUT = ROOT / "outputs/b-tube-cut"
STOCKS = [9000.0, 10000.0, 11000.0, 12000.0]
REMNANT_MIN = 200.0  # mm
GIDS = [f"G{i}" for i in range(1, 11)]
RESULT_TEMPLATES = tuple(f"result{i}.xlsx" for i in range(1, 5))
PROFILE_METHOD = "fixed_angular_neighborhood_conservative_v1"
# The sparsest supplied circumference has a roughly 12.05 degree maximum
# sample gap.  A fixed +/-6.25 degree physical neighbourhood covers every
# target angle at every output resolution.  Taking the most conservative end
# value in that neighbourhood prevents empty fine bins from creating fictitious
# co-cut spikes.
PROFILE_NEIGHBORHOOD_RADIUS_DEG = 6.25


def required_input_paths() -> list[Path]:
    """Files required for an actual Tube solve (STP files are not read)."""
    return (
        [CSV_DIR / f"圆管{i}.csv" for i in range(1, 11)]
        + [BATCH_XLSX]
        + [TMPL_DIR / name for name in RESULT_TEMPLATES]
    )


def input_readiness() -> dict:
    """Return a machine-readable preflight without touching solver outputs."""
    required = required_input_paths()
    missing = [p for p in required if not p.is_file()]

    def rel(path: Path) -> str:
        try:
            return path.relative_to(ROOT).as_posix()
        except ValueError:
            return path.as_posix()

    return {
        "ok": not missing,
        "problem_id": "tube_cut_b2026",
        "required_count": len(required),
        "missing_count": len(missing),
        "missing_inputs": [rel(p) for p in missing],
        "data_manifest": "fixtures/t3/tube_cut_b2026/DATA_REQUIRED.md",
    }


def input_sha256() -> dict[str, str]:
    """Hash every authorised source/template used by a real solve."""
    ready = input_readiness()
    if not ready["ok"]:
        raise FileNotFoundError("cannot hash Tube inputs before readiness passes")
    hashes: dict[str, str] = {}
    for path in required_input_paths():
        rel = path.relative_to(ROOT).as_posix()
        digest = hashlib.sha256()
        with path.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
        hashes[rel] = digest.hexdigest()
    return hashes


def blocked_envelope(problem_id: str = "tube_cut_b2026") -> dict:
    """Honest product response when the untracked contest attachments are absent."""
    ready = input_readiness()
    reason = (
        "required Tube source attachments are absent; restore the original local "
        "contest data listed in the data manifest before solving"
    )
    return {
        "problem_id": problem_id or "tube_cut_b2026",
        "problem_class": "tube_cut",
        "status": "BLOCKED",
        "objective": None,
        "source": "tools/solve_tube_cut_b2026.py",
        "solver": "none",
        "questions": {},
        "missing_inputs": ready["missing_inputs"],
        "data_manifest": ready["data_manifest"],
        "meta": {
            "exact": False,
            "proven_optimal": False,
            "method_class": "heuristic",
            "blocked": True,
            "blocked_code": "tube_source_data_missing",
            "reason": reason,
            "required_count": ready["required_count"],
            "missing_count": ready["missing_count"],
        },
    }


# ---------------------------------------------------------------------------
# Geometry
# ---------------------------------------------------------------------------

def load_points(i: int) -> np.ndarray:
    df = pd.read_csv(CSV_DIR / f"圆管{i}.csv")
    return df[["X", "Y", "Z"]].to_numpy(dtype=float)


def analyze_tube(i: int, *, n_bins: int = 360) -> dict:
    P = load_points(i)
    c = P.mean(axis=0)
    _, _, Vt = np.linalg.svd(P - c, full_matrices=False)
    axis = Vt[0].copy()
    # Deterministic PCA sign: the largest absolute component is positive.
    pivot = int(np.argmax(np.abs(axis)))
    if axis[pivot] < 0:
        axis = -axis
    t = (P - c) @ axis
    t_min, t_max = float(t.min()), float(t.max())
    length = t_max - t_min
    # orthonormal frame
    tmp = np.array([0.0, 0.0, 1.0]) if abs(axis[2]) < 0.9 else np.array([0.0, 1.0, 0.0])
    e1 = np.cross(axis, tmp)
    e1 /= np.linalg.norm(e1) + 1e-15
    e2 = np.cross(axis, e1)
    e2 /= np.linalg.norm(e2) + 1e-15
    # Full-circumference cut profiles, not a fixed 3 mm end band.
    L_env = end_envelope(P, c, axis, e1, e2, t, t_min, side="L", n_bins=n_bins)
    R_env = end_envelope(P, c, axis, e1, e2, t, t_max, side="R", n_bins=n_bins)
    return {
        "id": f"G{i}",
        "axial_length_mm": round(length, 4),
        "axis": axis.tolist(),
        "center": c.tolist(),
        "t_min": t_min,
        "t_max": t_max,
        "L_env": L_env.tolist(),
        "R_env": R_env.tolist(),
        "n_points": int(len(P)),
        "profile_bins": n_bins,
    }


def end_envelope(P, c, axis, e1, e2, t, t_end, *, side: str, n_bins: int) -> np.ndarray:
    """Return a conservative, resolution-stable end profile in millimetres.

    Every target ray uses a fixed physical angular neighbourhood rather than a
    bin whose width shrinks with ``n_bins``.  The L profile takes the minimum
    axial coordinate in the neighbourhood and the R profile takes the maximum;
    both choices reduce, never inflate, the inferred nesting inset.  The fixed
    radius covers the largest observed angular gap in the authorised point
    clouds and makes 180/360/720 resolution a rotation-search refinement, not a
    change in the source geometry.
    """
    rel = P - c
    ang = np.arctan2(rel @ e2, rel @ e1)  # [-pi,pi]
    targets = -np.pi + (np.arange(n_bins, dtype=float) + 0.5) * 2 * np.pi / n_bins
    radius = np.deg2rad(PROFILE_NEIGHBORHOOD_RADIUS_DEG)
    env = np.empty(n_bins, dtype=float)
    for index, target in enumerate(targets):
        distance = np.abs(np.angle(np.exp(1j * (ang - target))))
        selected = distance <= radius + 1e-12
        if not selected.any():
            # Defensive fallback for a different future data set.  The current
            # authorised inputs never need it because the radius covers their
            # maximum angular gap; using nearest samples remains conservative.
            selected = distance <= float(distance.min()) + 1e-12
        axial = float(np.min(t[selected]) if side == "L" else np.max(t[selected]))
        env[index] = axial - t_end if side == "L" else t_end - axial
    return np.maximum(env, 0.0)


def cocut_saving(geo_i: dict, geo_j: dict, end_i: str, end_j: str) -> float:
    """Δ from a full discrete rotation search of the two facing end profiles."""
    ei = np.array(geo_i[f"{end_i}_env"], dtype=float)
    ej = np.array(geo_j[f"{end_j}_env"], dtype=float)
    n = len(ei)
    best = 0.0
    # A proper 3-D placement fixes angular handedness.  With i on the left and
    # j on the right, equal end labels (LL/RR) reverse angular direction;
    # different labels (LR/RL) preserve it.  Trying both and taking max would
    # admit a mirror reflection that is not a physical rotation.
    facing = ej[::-1] if end_i == end_j else ej
    for shift in range(n):
        nest = float(np.min(ei + np.roll(facing, shift)))
        if nest > best:
            best = nest
    # Cap by physical lengths
    li, lj = geo_i["axial_length_mm"], geo_j["axial_length_mm"]
    best = min(best, li * 0.5, lj * 0.5, li + lj)
    return round(max(0.0, best), 4)


def build_geometry(*, n_bins: int = 360) -> dict:
    geos = {f"G{i}": analyze_tube(i, n_bins=n_bins) for i in range(1, 11)}
    lengths = {g: geos[g]["axial_length_mm"] for g in GIDS}
    # 10x10x4 savings
    modes = [("L", "L"), ("L", "R"), ("R", "L"), ("R", "R")]
    mode_names = ["LL", "LR", "RL", "RR"]
    savings: dict[str, dict[str, float]] = {}
    for a in GIDS:
        for b in GIDS:
            key = f"{a}-{b}"
            savings[key] = {
                mn: cocut_saving(geos[a], geos[b], ea, eb)
                for mn, (ea, eb) in zip(mode_names, modes)
            }
    return {
        "geos": geos,
        "lengths": lengths,
        "savings": savings,
        "profile_bins": n_bins,
        "profile_method": PROFILE_METHOD,
        "angular_neighborhood_radius_deg": PROFILE_NEIGHBORHOOD_RADIUS_DEG,
        "rotation_step_deg": 360.0 / n_bins,
    }


def solve_q1(
    lengths: dict,
    *,
    seed: int,
    time_limit_s: float,
    secondary_time_limit_s: float,
) -> dict:
    demand = {g: 50 for g in GIDS}
    primary_bins = mixed_stock_patterns(
        demand,
        lengths,
        stocks=STOCKS,
        seed=seed,
        time_limit_s=time_limit_s,
    )
    secondary_trials = []
    secondary_candidates = []
    for secondary_seed in dict.fromkeys([seed, seed + 17]):
        rows, evidence = minimize_type_splits_for_fixed_stock(
            demand,
            lengths,
            primary_bins,
            stocks=STOCKS,
            seed=secondary_seed,
            time_limit_s=secondary_time_limit_s,
        )
        secondary_trials.append(evidence)
        secondary_candidates.append((rows, evidence))
    bins, selected_secondary = min(
        secondary_candidates,
        key=lambda item: (
            int(item[1]["switch_incumbent"]),
            -int(item[1]["switch_lower_bound"]),
            int(item[1]["seed"]),
        ),
    )
    secondary = {
        "selected_seed": selected_secondary["seed"],
        "selected": selected_secondary,
        "trials": secondary_trials,
        # Duplicate the key totals at this level for validator/backward readers.
        "switch_incumbent": selected_secondary["switch_incumbent"],
        "switch_lower_bound": selected_secondary["switch_lower_bound"],
    }
    for b in bins:
        used = sum(float(lengths[g]) for g in b["sequence"])
        b.update(
            {
                "used_length_mm": round(used, 6),
                "raw_length_mm": round(used, 6),
                "effective_length_mm": round(used, 6),
                "co_cut_benefit_mm": 0.0,
                "leftover_mm": round(b["stock_length_mm"] - used, 6),
                "utilization": round(used / b["stock_length_mm"], 8),
                "switches": sum(
                    a != c for a, c in zip(b["sequence"], b["sequence"][1:])
                ),
                "orientations": [],
                "joints": [],
            }
        )
    total_stock = sum(b["stock_length_mm"] for b in bins)
    total_raw = sum(b["used_length_mm"] for b in bins)
    # With no co-cut in Q1, any stock plan must cover the raw axial length.
    primary_bound = joint_relaxation_stock_lower_bound(
        demand,
        lengths,
        {
            f"{a}-{b}": {mode: 0.0 for mode in ("LL", "LR", "RL", "RR")}
            for a in GIDS
            for b in GIDS
        },
        stocks=STOCKS,
    )
    primary_lb = float(primary_bound["lower_bound_mm"])
    return {
        "stocks": bins,
        "total_stock_length_mm": round(total_stock, 3),
        "total_axial_length_mm": round(total_raw, 4),
        "utilization": round(total_raw / total_stock, 6),
        "total_switch": sum(b["switches"] for b in bins),
        "status": "FEASIBLE",
        "exact": False,
        "method": "lexicographic compact CP-SAT: stock length, then type switches",
        "demand": demand,
        "seed": seed,
        "optimality": {
            "primary_lower_bound_mm": primary_lb,
            "primary_gap_mm": round(total_stock - primary_lb, 6),
            "primary_proven_optimal": abs(total_stock - primary_lb) <= 1e-6,
            "secondary": secondary,
        },
        "search_evidence": {
            "primary_incumbent_stock_mm": round(
                sum(float(row["stock_length_mm"]) for row in primary_bins), 6
            ),
            "primary_incumbent_switches": sum(
                sum(a != b for a, b in zip(row["sequence"], row["sequence"][1:]))
                for row in primary_bins
            ),
            "secondary_selected": True,
        },
        "_q3_alternative_base": primary_bins,
    }


def solve_q2(
    q1: dict,
    lengths: dict,
    savings: dict,
    *,
    seed: int,
    iterations: int,
    restarts: int = 1,
) -> dict:
    base = [
        {
            "id": s["id"],
            "stock_length_mm": s["stock_length_mm"],
            "purchase_cost_mm": s["stock_length_mm"],
            "sequence": list(s["sequence"]),
            "from_remnant": False,
        }
        for s in q1["stocks"]
    ]
    experiments = []
    candidates = []
    for restart in range(max(1, int(restarts))):
        restart_seed = seed + restart * 1009
        stocks = optimize_fixed_assignments(
            base,
            lengths,
            savings,
            seed=restart_seed,
            iterations=iterations,
        )
        candidate = summarize_bins(stocks)
        experiments.append(
            {
                "restart": restart,
                "seed": restart_seed,
                "total_co_cut_benefit_mm": candidate["total_co_cut_benefit_mm"],
                "total_switch": candidate["total_switch"],
            }
        )
        candidates.append((candidate, restart_seed))
    res, selected_seed = min(
        candidates,
        key=lambda item: (
            -float(item[0]["total_co_cut_benefit_mm"]),
            int(item[0]["total_switch"]),
            int(item[1]),
        ),
    )
    res["note"] = "Assignment fixed from Q1; reorder+joints only"
    res["method"] = "Q1 fixed assignment + multi-start seeded ALNS + exact orientation DP"
    res["demand"] = dict(q1["demand"])
    res["seed"] = selected_seed
    res["search_evidence"] = {
        "base_seed": seed,
        "selected_seed": selected_seed,
        "restarts": experiments,
    }
    return res


def solve_q3(
    q2: dict,
    lengths: dict,
    savings: dict,
    *,
    seed: int,
    iterations: int,
    restarts: int = 1,
    alternative_bases: list[list[dict]] | None = None,
    block_master_time_limit_s: float | None = None,
) -> dict:
    demand = dict(q2["demand"])
    bases = [("q2_low_switch", deepcopy(q2["stocks"]))]
    for index, rows in enumerate(alternative_bases or [], 1):
        if counts_from_bins(rows) != demand:
            raise ValueError(f"Q3 alternative base {index} does not match demand")
        bases.append((f"primary_packing_{index}", deepcopy(rows)))
    block_evidence = None
    if block_master_time_limit_s is not None and block_master_time_limit_s > 0:
        block_trials = []
        # The canonical seed is retained as a stable cross-run baseline; the
        # Q3-derived seeds add diversified deterministic starts.
        block_seeds = list(
            dict.fromkeys([20260813, seed + 7919, seed + 8928])
        )
        block_candidates = []
        for block_seed in block_seeds:
            rows, evidence = homogeneous_block_stock_patterns(
                demand,
                lengths,
                savings,
                stocks=STOCKS,
                seed=block_seed,
                time_limit_s=block_master_time_limit_s,
                id_prefix="M",
            )
            block_trials.append(evidence)
            block_candidates.append((rows, evidence))
        block_rows, selected_block_evidence = min(
            block_candidates,
            key=lambda item: (
                float(item[1]["stock_incumbent_mm"]),
                int(item[1]["type_incidence"]),
                int(item[1]["seed"]),
            ),
        )
        block_evidence = {
            "selected_seed": selected_block_evidence["seed"],
            "selected": selected_block_evidence,
            "trials": block_trials,
        }
        bases.append(("homogeneous_block_master", block_rows))
    experiments = []
    candidates = []
    if block_evidence is not None:
        block_candidate = summarize_bins(deepcopy(block_rows))
        experiments.append(
            {
                "restart": "block-baseline",
                "seed": block_evidence["selected_seed"],
                "base": "homogeneous_block_master",
                "total_stock_length_mm": block_candidate["total_stock_length_mm"],
                "total_co_cut_benefit_mm": block_candidate["total_co_cut_benefit_mm"],
                "total_switch": block_candidate["total_switch"],
            }
        )
        candidates.append((block_candidate, int(block_evidence["selected_seed"])))
    planned_restarts = max(max(1, int(restarts)), len(bases))
    for restart in range(planned_restarts):
        restart_seed = seed + restart * 1009
        base_name, base = bases[restart % len(bases)]
        stocks = optimize_joint_bins(
            base,
            lengths,
            savings,
            seed=restart_seed,
            iterations=iterations,
            resize_new=True,
            stocks=STOCKS,
        )
        if counts_from_bins(stocks) != demand:
            raise AssertionError("Q3 joint search changed demand")
        candidate = summarize_bins(stocks)
        experiments.append(
            {
                "restart": restart,
                "seed": restart_seed,
                "base": base_name,
                "total_stock_length_mm": candidate["total_stock_length_mm"],
                "total_co_cut_benefit_mm": candidate["total_co_cut_benefit_mm"],
                "total_switch": candidate["total_switch"],
            }
        )
        candidates.append((candidate, restart_seed))
    res, selected_seed = min(
        candidates,
        key=lambda item: (
            float(item[0]["total_stock_length_mm"]),
            -float(item[0]["total_co_cut_benefit_mm"]),
            int(item[0]["total_switch"]),
            int(item[1]),
        ),
    )
    bound = joint_relaxation_stock_lower_bound(
        demand, lengths, savings, stocks=STOCKS
    )
    lower = float(bound["lower_bound_mm"])
    incumbent = float(res["total_stock_length_mm"])
    res["method"] = "compact mixed-stock master + multi-start joint seeded ALNS + exact orientation DP"
    res["demand"] = demand
    res["seed"] = selected_seed
    res["optimality"] = {
        **bound,
        "incumbent_mm": incumbent,
        "absolute_gap_mm": round(incumbent - lower, 6),
        "relative_gap_to_lower_bound": round((incumbent - lower) / lower, 8)
        if lower
        else 0.0,
        "proven_optimal": abs(incumbent - lower) <= 1e-6,
    }
    res["search_evidence"] = {
        "base_seed": seed,
        "selected_seed": selected_seed,
        "restarts": experiments,
        "homogeneous_block_master": block_evidence,
    }
    return res


def load_batches() -> list[dict[str, int]]:
    df = pd.read_excel(BATCH_XLSX, header=None)
    # find header row
    batches = []
    # rows: 工件1..10 in col0, counts in col1-3
    for bi in range(3):
        dem = {}
        for i in range(1, 11):
            # search
            for r in range(len(df)):
                val = df.iloc[r, 0]
                if str(val).strip() in (f"工件{i}", f"G{i}", str(i)):
                    dem[f"G{i}"] = int(df.iloc[r, bi + 1])
                    break
        if len(dem) < 10:
            missing = sorted(set(GIDS) - set(dem))
            raise ValueError(
                f"batch workbook is missing demand rows for B{bi + 1}: {missing}; "
                "hard-coded OCR fallback is forbidden"
            )
        if any(value < 0 for value in dem.values()):
            raise ValueError(f"batch B{bi + 1} contains negative demand")
        batches.append(dem)
    return batches


def solve_q4(
    lengths: dict,
    savings: dict,
    *,
    seed: int,
    beam_width: int,
    iterations: int,
    master_time_limit_s: float,
    block_master_time_limit_s: float,
    remaining_block_master_time_limit_s: float,
    remaining_block_max_keys_per_batch: int,
    restarts: int = 1,
) -> dict:
    batches = load_batches()
    block_initial_bins: list[list[dict] | None] = []
    block_trials = []
    for batch_index, demand in enumerate(batches, 1):
        block_seed = seed + batch_index * 7919
        try:
            rows, evidence = homogeneous_block_stock_patterns(
                demand,
                lengths,
                savings,
                stocks=STOCKS,
                seed=block_seed,
                time_limit_s=block_master_time_limit_s,
                id_prefix=f"B{batch_index}-C",
            )
            block_initial_bins.append(rows)
            block_trials.append(
                {
                    "batch": f"B{batch_index}",
                    "available": True,
                    **evidence,
                }
            )
        except RuntimeError as exc:
            # The raw master remains a complete, feasible fallback.  Record a
            # short-budget UNKNOWN instead of failing or hiding the experiment.
            block_initial_bins.append(None)
            block_trials.append(
                {
                    "batch": f"B{batch_index}",
                    "available": False,
                    "seed": block_seed,
                    "time_limit_s": block_master_time_limit_s,
                    "reason": str(exc),
                }
            )
    experiments = []
    candidates = []
    for restart in range(max(1, int(restarts))):
        restart_seed = seed + restart * 1009
        candidate = solve_multibatch_beam(
            batches,
            lengths,
            savings,
            stocks=STOCKS,
            remnant_min_mm=REMNANT_MIN,
            seed=restart_seed,
            beam_width=beam_width,
            variants_per_state=6,
            joint_iterations=iterations,
            master_time_limit_s=master_time_limit_s,
            cocut_initial_bins=block_initial_bins,
            cocut_remaining_master_time_limit_s=remaining_block_master_time_limit_s,
            cocut_remaining_max_keys_per_batch=remaining_block_max_keys_per_batch,
        )
        experiments.append(
            {
                "restart": restart,
                "seed": restart_seed,
                "total_new_standard_stock_mm": candidate["total_new_standard_stock_mm"],
                "total_waste_mm": candidate["total_waste_mm"],
                "final_inventory_mm": candidate["final_inventory_mm"],
                "total_co_cut_benefit_mm": candidate["total_co_cut_benefit_mm"],
                "total_switch": candidate["total_switch"],
            }
        )
        candidates.append((candidate, restart_seed))
    result, selected_seed = min(
        candidates,
        key=lambda item: (
            float(item[0]["total_new_standard_stock_mm"]),
            -float(item[0]["total_co_cut_benefit_mm"]),
            int(item[0]["total_switch"]),
            float(item[0]["total_waste_mm"]),
            int(item[1]),
        ),
    )
    total_demand = {
        gid: sum(int(batch[gid]) for batch in batches) for gid in GIDS
    }
    bound = joint_relaxation_stock_lower_bound(
        total_demand, lengths, savings, stocks=STOCKS
    )
    lower = float(bound["lower_bound_mm"])
    incumbent = float(result["total_new_standard_stock_mm"])
    result["seed"] = selected_seed
    result["optimality"] = {
        **bound,
        "relaxations": [
            "all three batches pooled",
            "remnant timing and 200 mm threshold ignored",
            "path connectivity and per-bar capacity ignored before stock rounding",
        ],
        "incumbent_mm": incumbent,
        "absolute_gap_mm": round(incumbent - lower, 6),
        "relative_gap_to_lower_bound": round((incumbent - lower) / lower, 8)
        if lower
        else 0.0,
        "proven_optimal": abs(incumbent - lower) <= 1e-6,
    }
    result["search_evidence"] = {
        "base_seed": seed,
        "selected_seed": selected_seed,
        "restarts": experiments,
        "homogeneous_block_initializers": block_trials,
    }
    return result


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def sequence_blocks(sequence: list[str]) -> list[dict]:
    """Compress consecutive equal pieces into the contest's Gx×n blocks."""
    blocks: list[dict] = []
    for gid in sequence:
        if blocks and blocks[-1]["gid"] == gid:
            blocks[-1]["count"] += 1
        else:
            blocks.append({"gid": gid, "count": 1})
    for block in blocks:
        block["label"] = f"{block['gid']}×{block['count']}"
    return blocks


def sequence_block_text(sequence: list[str]) -> str:
    return "|".join(block["label"] for block in sequence_blocks(sequence))


def joint_summary(stock: dict) -> list[dict]:
    """Aggregate adjacent joints into the official internal/inter-block rows."""
    sequence = list(stock.get("sequence") or [])
    joints = list(stock.get("joints") or [])
    if len(joints) != max(0, len(sequence) - 1):
        raise ValueError(
            f"{stock.get('id')}: joints={len(joints)} does not match sequence={len(sequence)}"
        )
    if not sequence:
        return []
    blocks = sequence_blocks(sequence)
    position_blocks: list[str] = []
    for block in blocks:
        position_blocks.extend([block["label"]] * int(block["count"]))
    grouped: dict[tuple, dict] = {}
    for index, joint in enumerate(joints):
        front = position_blocks[index]
        rear = position_blocks[index + 1]
        kind = "内部拼接" if front == rear else "块间拼接"
        mode = str(joint["mode"])
        benefit = round(float(joint["benefit"]), 6)
        key = (kind, front, rear, mode, benefit)
        row = grouped.setdefault(
            key,
            {
                "type": kind,
                "front": front,
                "rear": rear,
                "mode": mode,
                "count": 0,
                "unit_benefit_mm": benefit,
                "subtotal_mm": 0.0,
            },
        )
        row["count"] += 1
        row["subtotal_mm"] = round(row["count"] * benefit, 6)
    return list(grouped.values())


def trace_q4_materials(q4: dict) -> tuple[list[dict], list[dict], list[dict]]:
    """Trace each Q4 remnant back to one purchased standard tube."""
    remnant_owner: dict[str, str] = {}
    materials: dict[str, dict] = {}
    plans: list[dict] = []
    joints: list[dict] = []
    next_material = 1
    for batch in q4["batches"]:
        batch_id = str(batch["batch"])
        result = batch["result"]
        next_remnant_owner: dict[str, str] = {
            str(row["id"]): remnant_owner[str(row["id"])]
            for row in result.get("inventory_before", [])
            if str(row["id"]) in remnant_owner
        }
        for row_index, stock in enumerate(result["stocks"], 1):
            if stock.get("from_remnant"):
                source_id = str(stock["remnant_id"])
                if source_id not in remnant_owner:
                    raise ValueError(f"unowned remnant in {batch_id}: {source_id}")
                material_id = remnant_owner[source_id]
                source = "余料"
                display_id = source_id
            else:
                material_id = f"M{next_material}"
                next_material += 1
                source = "新母材"
                display_id = material_id
                materials[material_id] = {
                    "material_id": material_id,
                    "first_batch": batch_id,
                    "last_batch": batch_id,
                    "initial_length_mm": float(stock["stock_length_mm"]),
                    "total_co_cut_benefit_mm": 0.0,
                    "total_effective_length_mm": 0.0,
                    "final_leftover_mm": float(stock["stock_length_mm"]),
                    "final_status": "入库",
                }
            material = materials[material_id]
            leftover = float(stock["leftover_mm"])
            if leftover <= 1e-6:
                status = "用尽"
            elif leftover >= REMNANT_MIN - 1e-6:
                status = "入库"
            else:
                status = "废料"
            material["last_batch"] = batch_id
            material["total_co_cut_benefit_mm"] = round(
                material["total_co_cut_benefit_mm"]
                + float(stock["co_cut_benefit_mm"]),
                6,
            )
            material["total_effective_length_mm"] = round(
                material["total_effective_length_mm"]
                + float(stock["effective_length_mm"]),
                6,
            )
            material["final_leftover_mm"] = round(leftover, 6)
            material["final_status"] = status
            plans.append(
                {
                    "batch": batch_id,
                    "display_id": display_id,
                    "source": source,
                    "stock_length_mm": float(stock["stock_length_mm"]),
                    "sequence": sequence_block_text(stock["sequence"]),
                    "co_cut_benefit_mm": float(stock["co_cut_benefit_mm"]),
                    "effective_length_mm": float(stock["effective_length_mm"]),
                    "leftover_mm": leftover,
                    "status": status,
                }
            )
            for summary in joint_summary(stock):
                joints.append(
                    {"batch": batch_id, "display_id": display_id, **summary}
                )
            if status == "入库":
                remnant_id = f"{batch_id}-R{row_index}"
                next_remnant_owner[remnant_id] = material_id
        inventory_ids = {str(row["id"]) for row in result.get("inventory_after", [])}
        remnant_owner = {
            remnant_id: owner
            for remnant_id, owner in next_remnant_owner.items()
            if remnant_id in inventory_ids
        }
    for material in materials.values():
        material["utilization"] = round(
            material["total_effective_length_mm"] / material["initial_length_mm"],
            8,
        )
    return plans, joints, list(materials.values())


def export_excels(lengths, savings, q1, q2, q3, q4):
    from openpyxl import load_workbook

    OUT.mkdir(parents=True, exist_ok=True)

    def copy_tmpl(name: str) -> Path:
        src = TMPL_DIR / name
        dst = OUT / name
        shutil.copy2(src, dst)
        return dst

    def clear_values(ws, start_row: int, end_row: int | None = None) -> None:
        for row in ws.iter_rows(
            min_row=start_row, max_row=end_row or ws.max_row, max_col=ws.max_column
        ):
            for cell in row:
                cell.value = None

    def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
        if source_row == target_row:
            return
        for column in range(1, max_col + 1):
            source = ws.cell(source_row, column)
            target = ws.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
        if ws.row_dimensions[source_row].height is not None:
            ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    def write_rows(ws, start_row: int, rows: list[list], *, style_row: int) -> None:
        clear_values(ws, start_row)
        max_col = max((len(row) for row in rows), default=ws.max_column)
        for offset, values in enumerate(rows):
            row_number = start_row + offset
            copy_row_style(ws, style_row, row_number, max_col)
            for column, value in enumerate(values, 1):
                ws.cell(row_number, column, value)

    def plan_rows(stocks: list[dict]) -> list[list]:
        return [
            [
                stock["id"],
                stock["stock_length_mm"],
                sequence_block_text(stock["sequence"]),
                stock["co_cut_benefit_mm"],
                stock["effective_length_mm"],
                stock["leftover_mm"],
                stock["utilization"],
            ]
            for stock in stocks
        ]

    def joint_rows(stocks: list[dict], include_batch: str | None = None) -> list[list]:
        rows: list[list] = []
        for stock in stocks:
            for summary in joint_summary(stock):
                prefix = [include_batch, stock["id"]] if include_batch else [stock["id"]]
                rows.append(
                    prefix
                    + [
                        summary["type"],
                        summary["front"],
                        summary["rear"],
                        summary["mode"],
                        summary["count"],
                        summary["unit_benefit_mm"],
                        summary["subtotal_mm"],
                    ]
                )
        return rows

    # --- result1 ---
    p1 = copy_tmpl("result1.xlsx")
    wb = load_workbook(p1)
    ws = wb["问题一_轴向占用长度"]
    for i, g in enumerate(GIDS, start=2):
        ws.cell(i, 1, g)
        ws.cell(i, 2, lengths[g])
    ws2 = wb["问题一_下料方案"]
    write_rows(
        ws2,
        2,
        [
            [
                s["id"],
                s["stock_length_mm"],
                sequence_block_text(s["sequence"]),
                s["used_length_mm"],
                s["leftover_mm"],
                s["utilization"],
            ]
            for s in q1["stocks"]
        ],
        style_row=2,
    )
    ws3 = wb["问题一_汇总指标"]
    ws3.cell(2, 1, q1["total_stock_length_mm"])
    ws3.cell(2, 2, q1["total_axial_length_mm"])
    ws3.cell(2, 3, q1["utilization"])
    ws3.cell(2, 4, q1["total_switch"])
    wb.save(p1)

    # --- result2 ---
    p2 = copy_tmpl("result2.xlsx")
    wb = load_workbook(p2)
    for mode in ("LL", "LR", "RL", "RR"):
        ws = wb[f"问题二_{mode}收益矩阵"]
        for i, left in enumerate(GIDS, 3):
            for j, right in enumerate(GIDS, 2):
                ws.cell(i, j, savings[f"{left}-{right}"][mode])
    ws = wb["问题二_下料方案"]
    write_rows(ws, 3, plan_rows(q2["stocks"]), style_row=3)
    ws = wb["问题二_拼接方式摘要表"]
    headers = [
        "母材编号(M_ID)", "拼接类型", "前工件块", "后工件块",
        "拼接方式", "拼接次数", "单次共切收益(mm)", "共切收益小计(mm)",
    ]
    for column, value in enumerate(headers, 1):
        ws.cell(2, column, value)
    write_rows(ws, 3, joint_rows(q2["stocks"]), style_row=3)
    ws = wb["问题二_汇总指标"]
    for column, value in enumerate(
        [
            q2["total_stock_length_mm"],
            q2["total_co_cut_benefit_mm"],
            q2["total_effective_length_mm"],
            q2["utilization"],
            q2["total_switch"],
        ],
        1,
    ):
        ws.cell(2, column, value)
    wb.save(p2)

    p3 = copy_tmpl("result3.xlsx")
    wb = load_workbook(p3)
    write_rows(wb["问题三_下料方案"], 2, plan_rows(q3["stocks"]), style_row=2)
    write_rows(
        wb["问题三_拼接方式摘要表"],
        2,
        joint_rows(q3["stocks"]),
        style_row=2,
    )
    ws = wb["问题三_汇总指标"]
    for column, value in enumerate(
        [
            q3["total_stock_length_mm"],
            q3["total_co_cut_benefit_mm"],
            q3["total_effective_length_mm"],
            q3["utilization"],
            q3["total_switch"],
        ],
        1,
    ):
        ws.cell(2, column, value)
    wb.save(p3)

    p4 = copy_tmpl("result4.xlsx")
    wb = load_workbook(p4)
    q4_plans, q4_joints, q4_materials = trace_q4_materials(q4)
    write_rows(
        wb["问题四_下料方案"],
        3,
        [
            [
                row["batch"], row["display_id"], row["source"],
                row["stock_length_mm"], row["sequence"],
                row["co_cut_benefit_mm"], row["effective_length_mm"],
                row["leftover_mm"], row["status"],
            ]
            for row in q4_plans
        ],
        style_row=3,
    )
    write_rows(
        wb["问题四_拼接方式摘要表"],
        2,
        [
            [
                row["batch"], row["display_id"], row["type"], row["front"],
                row["rear"], row["mode"], row["count"],
                row["unit_benefit_mm"], row["subtotal_mm"],
            ]
            for row in q4_joints
        ],
        style_row=2,
    )
    ws = wb["问题四_母材利用率"]
    note = "说明：本表仅统计各根新母材在整个周期结束后的最终利用率，不单独统计余料利用率。"
    write_rows(
        ws,
        3,
        [
            [
                row["material_id"], row["first_batch"], row["last_batch"],
                row["initial_length_mm"], row["total_co_cut_benefit_mm"],
                row["total_effective_length_mm"], row["final_leftover_mm"],
                row["final_status"], f"=F{3 + index}/D{3 + index}",
            ]
            for index, row in enumerate(q4_materials)
        ],
        style_row=3,
    )
    ws.cell(4 + len(q4_materials), 1, note)
    total_effective_q4 = round(
        sum(
            float(batch["result"]["total_effective_length_mm"])
            for batch in q4["batches"]
        ),
        6,
    )
    ws = wb["问题四_汇总指标"]
    for column, value in enumerate(
        [
            q4["total_new_standard_stock_mm"],
            q4["total_co_cut_benefit_mm"],
            total_effective_q4,
            total_effective_q4 / q4["total_new_standard_stock_mm"],
            q4["total_switch"],
        ],
        1,
    ):
        ws.cell(2, column, value)
    # No spreadsheet engine is required to generate the deliverable. Ask
    # Excel/LibreOffice to refresh the utilisation formulas on first open.
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcOnSave = True
    wb.save(p4)
    return [p1, p2, p3, p4]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Solve Tube B with reproducible search")
    parser.add_argument("--seed", type=int, default=20260813)
    parser.add_argument("--profile-bins", type=int, default=360)
    parser.add_argument("--master-time-limit-s", type=float, default=10.0)
    parser.add_argument("--q1-secondary-time-limit-s", type=float, default=30.0)
    parser.add_argument("--q3-block-time-limit-s", type=float, default=10.0)
    parser.add_argument("--q4-block-time-limit-s", type=float, default=10.0)
    parser.add_argument(
        "--q4-remaining-block-time-limit-s", type=float, default=2.0
    )
    parser.add_argument("--q4-remaining-block-max-keys", type=int, default=8)
    parser.add_argument("--q2-iterations", type=int, default=1500)
    parser.add_argument("--q3-iterations", type=int, default=5000)
    parser.add_argument("--q4-iterations", type=int, default=500)
    parser.add_argument("--q2-restarts", type=int, default=3)
    parser.add_argument("--q3-restarts", type=int, default=4)
    parser.add_argument("--q4-restarts", type=int, default=4)
    parser.add_argument("--beam-width", type=int, default=12)
    parser.add_argument(
        "--fast",
        action="store_true",
        help="short deterministic smoke budget (does not change model semantics)",
    )
    parser.add_argument(
        "--quality",
        action="store_true",
        help="larger search budget for final experiments",
    )
    args = parser.parse_args(argv)
    if args.fast and args.quality:
        parser.error("--fast and --quality are mutually exclusive")
    if args.profile_bins < 36 or args.profile_bins > 1440:
        parser.error("--profile-bins must be between 36 and 1440")
    if min(
        args.master_time_limit_s,
        args.q1_secondary_time_limit_s,
        args.q3_block_time_limit_s,
        args.q4_block_time_limit_s,
        args.q4_remaining_block_time_limit_s,
    ) <= 0:
        parser.error("solver time limits must be positive")
    if min(args.q2_iterations, args.q3_iterations, args.q4_iterations) < 0:
        parser.error("iteration counts must be non-negative")
    if args.beam_width < 1:
        parser.error("--beam-width must be positive")
    if args.q4_remaining_block_max_keys < 0:
        parser.error("--q4-remaining-block-max-keys must be non-negative")
    if min(args.q2_restarts, args.q3_restarts, args.q4_restarts) < 1:
        parser.error("restart counts must be positive")
    if args.fast:
        args.profile_bins = min(args.profile_bins, 180)
        args.master_time_limit_s = min(args.master_time_limit_s, 2.0)
        args.q2_iterations = min(args.q2_iterations, 250)
        args.q3_iterations = min(args.q3_iterations, 600)
        args.q4_iterations = min(args.q4_iterations, 80)
        args.beam_width = min(args.beam_width, 5)
        args.q1_secondary_time_limit_s = min(args.q1_secondary_time_limit_s, 10.0)
        args.q3_block_time_limit_s = min(args.q3_block_time_limit_s, 10.0)
        args.q4_block_time_limit_s = min(args.q4_block_time_limit_s, 5.0)
        args.q4_remaining_block_time_limit_s = min(
            args.q4_remaining_block_time_limit_s, 1.0
        )
        args.q4_remaining_block_max_keys = min(
            args.q4_remaining_block_max_keys, 8
        )
        args.q2_restarts = min(args.q2_restarts, 2)
        args.q3_restarts = min(args.q3_restarts, 2)
        args.q4_restarts = min(args.q4_restarts, 2)
    elif args.quality:
        args.profile_bins = max(args.profile_bins, 720)
        args.master_time_limit_s = max(args.master_time_limit_s, 30.0)
        args.q2_iterations = max(args.q2_iterations, 10000)
        args.q3_iterations = max(args.q3_iterations, 30000)
        args.q4_iterations = max(args.q4_iterations, 1500)
        args.beam_width = max(args.beam_width, 24)
        args.q1_secondary_time_limit_s = max(args.q1_secondary_time_limit_s, 120.0)
        args.q3_block_time_limit_s = max(args.q3_block_time_limit_s, 30.0)
        args.q4_block_time_limit_s = max(args.q4_block_time_limit_s, 30.0)
        args.q4_remaining_block_time_limit_s = max(
            args.q4_remaining_block_time_limit_s, 3.0
        )
        args.q4_remaining_block_max_keys = max(
            args.q4_remaining_block_max_keys, 12
        )
        args.q2_restarts = max(args.q2_restarts, 8)
        args.q3_restarts = max(args.q3_restarts, 12)
        args.q4_restarts = max(args.q4_restarts, 12)

    ready = input_readiness()
    if not ready["ok"]:
        print(json.dumps(blocked_envelope(), ensure_ascii=False, indent=2))
        return 0
    OUT.mkdir(parents=True, exist_ok=True)
    print("Building geometry...", flush=True)
    geo = build_geometry(n_bins=args.profile_bins)
    lengths = geo["lengths"]
    savings = geo["savings"]
    (OUT / "axial_lengths.json").write_text(
        json.dumps(
            {
                "source": "PCA first principal axis span of point cloud (tube axis)",
                "bugfix": "Pi used Z_max-Z_min (~40mm cross-section); corrected to axial PCA length",
                "axial_lengths_mm": lengths,
                "total_one_each_mm": round(sum(lengths.values()), 4),
                "total_50_each_mm": round(50 * sum(lengths.values()), 4),
            },
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    (OUT / "tube_geometry.json").write_text(
        json.dumps({k: {kk: vv for kk, vv in g.items() if kk not in ("L_env", "R_env")} for k, g in geo["geos"].items()}, indent=2),
        encoding="utf-8",
    )
    (OUT / "cocut_savings.json").write_text(
        json.dumps(savings, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (OUT / "model_snapshot.json").write_text(
        json.dumps(
            {
                "schema": "orpath.tube_model.v2",
                "units": "mm",
                "lengths": lengths,
                "savings": savings,
                "stock_lengths_mm": STOCKS,
                "remnant_min_mm": REMNANT_MIN,
                "profile_bins": geo["profile_bins"],
                "profile_method": geo["profile_method"],
                "angular_neighborhood_radius_deg": geo[
                    "angular_neighborhood_radius_deg"
                ],
                "rotation_step_deg": geo["rotation_step_deg"],
                "seed": args.seed,
                "search_settings": {
                    "master_time_limit_s": args.master_time_limit_s,
                    "q1_secondary_time_limit_s": args.q1_secondary_time_limit_s,
                    "q3_block_time_limit_s": args.q3_block_time_limit_s,
                    "q4_block_time_limit_s": args.q4_block_time_limit_s,
                    "q4_remaining_block_time_limit_s": args.q4_remaining_block_time_limit_s,
                    "q4_remaining_block_max_keys": args.q4_remaining_block_max_keys,
                    "q2_iterations": args.q2_iterations,
                    "q3_iterations": args.q3_iterations,
                    "q4_iterations": args.q4_iterations,
                    "q2_restarts": args.q2_restarts,
                    "q3_restarts": args.q3_restarts,
                    "q4_restarts": args.q4_restarts,
                    "beam_width": args.beam_width,
                },
                "collaboration_protocol": {
                    "schema": "orpath.tube_collaboration.v1",
                    "budget_percent": {
                        "geometry": 20,
                        "q1_q2": 10,
                        "q3": 25,
                        "q4": 45,
                    },
                    "numeric_authority": "solve_plus_independent_validate",
                },
                "input_sha256": input_sha256(),
            },
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )

    print("lengths", lengths, flush=True)
    print("Q1...", flush=True)
    q1 = solve_q1(
        lengths,
        seed=args.seed,
        time_limit_s=args.master_time_limit_s,
        secondary_time_limit_s=args.q1_secondary_time_limit_s,
    )
    print("Q1 total", q1["total_stock_length_mm"], "sw", q1["total_switch"], flush=True)
    print("Q2...", flush=True)
    q2 = solve_q2(
        q1,
        lengths,
        savings,
        seed=args.seed + 101,
        iterations=args.q2_iterations,
        restarts=args.q2_restarts,
    )
    print("Q2 cocut", q2["total_co_cut_benefit_mm"], flush=True)
    print("Q3...", flush=True)
    q3_alternative_base = q1.pop("_q3_alternative_base")
    q3 = solve_q3(
        q2,
        lengths,
        savings,
        seed=args.seed + 202,
        iterations=args.q3_iterations,
        restarts=args.q3_restarts,
        alternative_bases=[q3_alternative_base],
        block_master_time_limit_s=args.q3_block_time_limit_s,
    )
    print("Q3 total", q3["total_stock_length_mm"], "cocut", q3["total_co_cut_benefit_mm"], flush=True)
    # keep better of q3 vs q1+q2 packing on primary then secondary
    if (q3["total_stock_length_mm"], -q3["total_co_cut_benefit_mm"], q3["total_switch"]) > (
        q1["total_stock_length_mm"],
        -q2["total_co_cut_benefit_mm"],
        q2["total_switch"],
    ):
        # q3 worse on primary: fallback to q2 plan as q3 baseline then try improve stocks
        print("Q3 not better on primary; using Q1 pack + cocut as Q3", flush=True)
        q3 = deepcopy(q2)
        q3["note"] = "Q2 fixed-assignment plan retained because joint search was not lexicographically better"
    print("Q4...", flush=True)
    q4 = solve_q4(
        lengths,
        savings,
        seed=args.seed + 303,
        beam_width=args.beam_width,
        iterations=args.q4_iterations,
        master_time_limit_s=args.master_time_limit_s,
        block_master_time_limit_s=args.q4_block_time_limit_s,
        remaining_block_master_time_limit_s=args.q4_remaining_block_time_limit_s,
        remaining_block_max_keys_per_batch=args.q4_remaining_block_max_keys,
        restarts=args.q4_restarts,
    )
    print("Q4 new stock", q4["total_new_standard_stock_mm"], flush=True)

    for name, obj in [("q1", q1), ("q2", q2), ("q3", q3), ("q4", q4)]:
        (OUT / f"{name}-solution.json").write_text(
            json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )

    print("Excel...", flush=True)
    paths = export_excels(lengths, savings, q1, q2, q3, q4)

    q1_primary_status = (
        "主目标已证明最优"
        if q1["optimality"]["primary_proven_optimal"]
        else "主目标尚未证明最优"
    )
    q3_primary_status = (
        "主目标已证明最优"
        if q3.get("optimality", {}).get("proven_optimal")
        else "主目标尚未证明最优"
    )
    q4_gap = q4["optimality"]
    done = f"""# DONE — 异形圆管下料可复现实验

## Bug fixes
1. **轴向长度**: PCA 第一主轴跨度（管轴），不再用 Z 截面 ~40mm  
2. **共切**: 固定 ±{geo['angular_neighborhood_radius_deg']}° 邻域的保守端部轮廓，{geo['profile_bins']} 个角度 + 全旋转搜索，步长 {geo['rotation_step_deg']}°
3. **数字单一来源**: 下列指标均来自 `q*-solution.json`  
4. **Excel**: 已重写 `result1.xlsx`…`result4.xlsx`
5. **复现参数**: seed={args.seed}; q1-secondary={args.q1_secondary_time_limit_s}s; q3-block={args.q3_block_time_limit_s}s; q4-block={args.q4_block_time_limit_s}s; q4-remaining-block={args.q4_remaining_block_time_limit_s}s/每批最多{args.q4_remaining_block_max_keys}个; q2={args.q2_iterations}x{args.q2_restarts}; q3={args.q3_iterations}x{args.q3_restarts}; q4={args.q4_iterations}x{args.q4_restarts}; beam={args.beam_width}

## 轴向长度 (mm)
{json.dumps(lengths, ensure_ascii=False)}

## 结果总表
| 问 | 总母材长度 mm | 总共切 mm | 切换 | 利用率/备注 |
|----|---------------|-----------|------|-------------|
| Q1 | {q1['total_stock_length_mm']} | 0 | {q1['total_switch']} | util={q1['utilization']} |
| Q2 | {q2['total_stock_length_mm']} | {q2['total_co_cut_benefit_mm']} | {q2['total_switch']} | util={q2['utilization']} 分配同Q1 |
| Q3 | {q3['total_stock_length_mm']} | {q3['total_co_cut_benefit_mm']} | {q3['total_switch']} | util={q3['utilization']} |
| Q4 | {q4['total_new_standard_stock_mm']} | {q4['total_co_cut_benefit_mm']} | {q4['total_switch']} | 新标准母材；余料≥200复用 |

## 状态
- Q1：{q1_primary_status}；固定母材总长后的切换数尚未证明最优。
- Q3：{q3_primary_status}（下界 {q3.get('optimality', {}).get('lower_bound_mm')} mm）。
- Q4：严格可行但尚未证明全局最优；下界 {q4_gap['lower_bound_mm']} mm，绝对差 {q4_gap['absolute_gap_mm']} mm，相对差 {100 * q4_gap['relative_gap_to_lower_bound']:.2f}%。
- 共切为高分辨率点云轮廓模型；仍不是 CAD 实体布尔碰撞证明

## 文件
- {paths[0]}
- {paths[1]}
- {paths[2]}
- {paths[3]}
- q1-solution.json … q4-solution.json
"""
    (OUT / "DONE.md").write_text(done, encoding="utf-8")
    (OUT / "MONITOR.md").write_text(
        "# MONITOR\n\nReproducible mixed-stock/ALNS/beam run complete. See DONE.md.\n",
        encoding="utf-8",
    )
    print("DONE", flush=True)
    print(done)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
